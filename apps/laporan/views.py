# Copyright (c) 2026 Dafa Al Hafiz. All rights reserved.
from django.shortcuts import render

# Create your views here.
from django.contrib.auth.mixins import LoginRequiredMixin
from django.views.generic import TemplateView
from django.http import HttpResponse
from django.utils import timezone
from .utils import get_rekap_periode, get_rekap_per_kapal
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

class LaporanIndexView(LoginRequiredMixin, TemplateView):
    template_name = 'laporan/index.html'

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        now = timezone.now()
        bulan = int(self.request.GET.get('bulan', now.month))
        tahun = int(self.request.GET.get('tahun', now.year))

        ctx['rekap'] = get_rekap_periode(bulan, tahun)
        ctx['rekap_kapal'] = get_rekap_per_kapal(tahun)
        ctx['bulan'] = bulan
        ctx['tahun'] = tahun
        ctx['bulan_list'] = range(1, 13)
        ctx['tahun_list'] = range(2023, now.year + 1)
        ctx['nama_bulan'] = [
            '', 'Januari', 'Februari', 'Maret', 'April', 'Mei', 'Juni',
            'Juli', 'Agustus', 'September', 'Oktober', 'November', 'Desember'
        ]
        return ctx

class ExportExcelView(LoginRequiredMixin, TemplateView):

    def get(self, request, *args, **kwargs):
        now = timezone.now()
        bulan = int(request.GET.get('bulan', now.month))
        tahun = int(request.GET.get('tahun', now.year))

        rekap = get_rekap_periode(bulan, tahun)

        # Buat workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Laporan Trip'

        # Style
        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill(fill_type='solid', fgColor='1a56db')
        center = Alignment(horizontal='center')

        nama_bulan = [
            '', 'Januari', 'Februari', 'Maret', 'April', 'Mei', 'Juni',
            'Juli', 'Agustus', 'September', 'Oktober', 'November', 'Desember'
        ]

        # Judul
        ws.merge_cells('A1:I1')
        ws['A1'] = f'Laporan Trip Usaha Bagan — {nama_bulan[bulan]} {tahun}'
        ws['A1'].font = Font(bold=True, size=13)
        ws['A1'].alignment = center

        # Header tabel
        headers = ['No', 'Kapal', 'Tgl Berangkat', 'Tgl Kembali',
                   'Total Tangkap (kg)', 'Pendapatan (Rp)',
                   'Biaya (Rp)', 'Bagi Hasil (Rp)', 'Laba Bersih (Rp)']

        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center

        # Data
        for i, d in enumerate(rekap['data'], 1):
            trip = d['trip']
            row = i + 3
            ws.cell(row=row, column=1, value=i)
            ws.cell(row=row, column=2, value=str(trip.kapal))
            ws.cell(row=row, column=3, value=str(trip.tgl_berangkat))
            ws.cell(row=row, column=4, value=str(trip.tgl_kembali or '-'))
            ws.cell(row=row, column=5, value=float(d['total_tangkap']))
            ws.cell(row=row, column=6, value=float(d['pendapatan']))
            ws.cell(row=row, column=7, value=float(d['biaya']))
            ws.cell(row=row, column=8, value=float(d['bagi_hasil']))
            ws.cell(row=row, column=9, value=float(d['laba']))

        # Baris total
        total_row = len(rekap['data']) + 4
        ws.cell(row=total_row, column=1, value='TOTAL').font = Font(bold=True)
        ws.cell(row=total_row, column=5, value=float(rekap['total_tangkap'])).font = Font(bold=True)
        ws.cell(row=total_row, column=6, value=float(rekap['total_pendapatan'])).font = Font(bold=True)
        ws.cell(row=total_row, column=7, value=float(rekap['total_biaya'])).font = Font(bold=True)
        ws.cell(row=total_row, column=8, value=float(rekap['total_bagi_hasil'])).font = Font(bold=True)
        ws.cell(row=total_row, column=9, value=float(rekap['total_laba'])).font = Font(bold=True)

        # Auto width kolom
        for col_cells in ws.iter_cols(min_row=1, max_row=ws.max_row):
            max_len = 0
            col_letter = None
            for cell in col_cells:
                if cell.value is not None and not isinstance(cell, openpyxl.cell.cell.MergedCell):
                    if col_letter is None:
                        col_letter = cell.column_letter
                    max_len = max(max_len, len(str(cell.value)))
            if col_letter:
                ws.column_dimensions[col_letter].width = max_len + 4

        # Response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename=laporan_{bulan}_{tahun}.xlsx'
        wb.save(response)
        return response