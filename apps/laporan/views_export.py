# Copyright (c) 2026 Dafa Al Hafiz. All rights reserved.
from django.http import HttpResponse
from django.shortcuts import get_object_or_404
from apps.core.mixins import OwnerRequiredMixin
from django.views.generic import TemplateView
from apps.operasional.models import Trip
from apps.tangkap.models import HasilTangkap
from apps.penjualan.models import Penjualan, BagiHasil
from .utils import get_rekap_periode, get_rekap_trip
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

class ExportExcelView(OwnerRequiredMixin, TemplateView):

    def get(self, request, *args, **kwargs):
        now = timezone.now()
        bulan = int(request.GET.get('bulan', now.month))
        tahun = int(request.GET.get('tahun', now.year))

        rekap = get_rekap_periode(bulan, tahun)

        # Buat workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Laporan Trip'

        # Style
        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill(fill_type='solid', fgColor='1a56db')
        center = Alignment(horizontal='center')

        nama_bulan = [
            '', 'Januari', 'Februari', 'Maret', 'April', 'Mei', 'Juni',
            'Juli', 'Agustus', 'September', 'Oktober', 'November', 'Desember'
        ]

        # Judul
        ws.merge_cells('A1:I1')
        ws['A1'] = f'Laporan Trip Usaha Bagan — {nama_bulan[bulan]} {tahun}'
        ws['A1'].font = Font(bold=True, size=13)
        ws['A1'].alignment = center

        # Header tabel
        headers = ['No', 'Kapal', 'Tgl Berangkat', 'Tgl Kembali',
                   'Total Tangkap (kg)', 'Pendapatan (Rp)',
                   'Biaya (Rp)', 'Bagi Hasil (Rp)', 'Laba Bersih (Rp)']

        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center

        # Data
        for i, d in enumerate(rekap['data'], 1):
            trip = d['trip']
            row = i + 3
            ws.cell(row=row, column=1, value=i)
            ws.cell(row=row, column=2, value=str(trip.kapal))
            ws.cell(row=row, column=3, value=str(trip.tgl_berangkat))
            ws.cell(row=row, column=4, value=str(trip.tgl_kembali or '-'))
            ws.cell(row=row, column=5, value=float(d['total_tangkap']))
            ws.cell(row=row, column=6, value=float(d['pendapatan']))
            ws.cell(row=row, column=7, value=float(d['biaya']))
            ws.cell(row=row, column=8, value=float(d['bagi_hasil']))
            ws.cell(row=row, column=9, value=float(d['laba']))

        # Baris total
        total_row = len(rekap['data']) + 4
        ws.cell(row=total_row, column=1, value='TOTAL').font = Font(bold=True)
        ws.cell(row=total_row, column=5, value=float(rekap['total_tangkap'])).font = Font(bold=True)
        ws.cell(row=total_row, column=6, value=float(rekap['total_pendapatan'])).font = Font(bold=True)
        ws.cell(row=total_row, column=7, value=float(rekap['total_biaya'])).font = Font(bold=True)
        ws.cell(row=total_row, column=8, value=float(rekap['total_bagi_hasil'])).font = Font(bold=True)
        ws.cell(row=total_row, column=9, value=float(rekap['total_laba'])).font = Font(bold=True)

        # Auto width kolom
        for col_cells in ws.iter_cols(min_row=1, max_row=ws.max_row):
            max_len = 0
            col_letter = None
            for cell in col_cells:
                if cell.value is not None and not isinstance(cell, openpyxl.cell.cell.MergedCell):
                    if col_letter is None:
                        col_letter = cell.column_letter
                    max_len = max(max_len, len(str(cell.value)))
            if col_letter:
                ws.column_dimensions[col_letter].width = max_len + 4

        # Response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename=laporan_{bulan}_{tahun}.xlsx'
        wb.save(response)
        return response

class ExportPDFTripView(OwnerRequiredMixin, TemplateView):
    """Ekspor laporan satu trip ke PDF menggunakan ReportLab."""

    def get(self, request, pk, *args, **kwargs):
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.units import cm
        from reportlab.lib import colors
        from reportlab.platypus import (
            SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, HRFlowable
        )
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.enums import TA_CENTER, TA_RIGHT

        trip = get_object_or_404(Trip.objects.select_related('kapal'), pk=pk)
        rekap = get_rekap_trip(trip)
        tangkap_list = HasilTangkap.objects.filter(trip=trip).select_related('jenis_ikan')
        jual_list = Penjualan.objects.filter(
            hasil_tangkap__trip=trip
        ).select_related('pembeli', 'hasil_tangkap__jenis_ikan')
        bagi_list = BagiHasil.objects.filter(trip=trip).select_related('abk')

        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename=laporan_trip_{trip.id}.pdf'

        doc = SimpleDocTemplate(response, pagesize=A4,
                                topMargin=2*cm, bottomMargin=2*cm,
                                leftMargin=2.5*cm, rightMargin=2.5*cm)

        styles = getSampleStyleSheet()
        h1 = ParagraphStyle('h1', parent=styles['Heading1'], fontSize=14,
                            spaceAfter=4, alignment=TA_CENTER)
        h2 = ParagraphStyle('h2', parent=styles['Heading2'], fontSize=11,
                            spaceBefore=12, spaceAfter=4)
        normal = ParagraphStyle('normal', parent=styles['Normal'], fontSize=9,
                                leading=14)
        right  = ParagraphStyle('right', parent=normal, alignment=TA_RIGHT)

        def rp(v):
            v = float(v or 0)
            if abs(v) >= 1e9: return f"Rp {v/1e9:.1f} M"
            if abs(v) >= 1e6: return f"Rp {v/1e6:.1f} jt"
            if abs(v) >= 1e3: return f"Rp {int(v/1e3)} rb"
            return f"Rp {int(v):,}"

        tbl_style = TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#1a56db')),
            ('TEXTCOLOR',  (0,0), (-1,0), colors.white),
            ('FONTNAME',   (0,0), (-1,0), 'Helvetica-Bold'),
            ('FONTSIZE',   (0,0), (-1,-1), 8),
            ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, colors.HexColor('#F8F9FA')]),
            ('GRID', (0,0), (-1,-1), 0.4, colors.HexColor('#DEE2E6')),
            ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
            ('TOPPADDING',  (0,0), (-1,-1), 4),
            ('BOTTOMPADDING', (0,0), (-1,-1), 4),
        ])

        story = []

        # Header
        story.append(Paragraph('Laporan Trip Usaha Bagan', h1))
        story.append(HRFlowable(width='100%', thickness=1, color=colors.HexColor('#1a56db')))
        story.append(Spacer(1, 0.3*cm))

        # Info trip
        info = [
            ['Kapal', trip.kapal.nama_kapal],
            ['Tgl Berangkat', str(trip.tgl_berangkat)],
            ['Tgl Kembali', str(trip.tgl_kembali or '-')],
            ['Status', trip.get_status_display()],
        ]
        t = Table(info, colWidths=[4*cm, 10*cm])
        t.setStyle(TableStyle([
            ('FONTSIZE', (0,0), (-1,-1), 9),
            ('FONTNAME', (0,0), (0,-1), 'Helvetica-Bold'),
            ('BOTTOMPADDING', (0,0), (-1,-1), 3),
        ]))
        story.append(t)
        story.append(Spacer(1, 0.4*cm))

        # Ringkasan keuangan
        story.append(Paragraph('Ringkasan Keuangan', h2))
        keu = [
            ['Total Tangkap', f"{float(rekap['total_tangkap']):.1f} kg"],
            ['Total Pendapatan', rp(rekap['pendapatan'])],
            ['Total Biaya Operasional', rp(rekap['biaya'])],
            ['Total Bagi Hasil ABK', rp(rekap['bagi_hasil'])],
            ['Laba Bersih', rp(rekap['laba'])],
        ]
        t = Table(keu, colWidths=[7*cm, 7*cm])
        t.setStyle(TableStyle([
            ('FONTSIZE', (0,0), (-1,-1), 9),
            ('FONTNAME', (0,4), (-1,4), 'Helvetica-Bold'),
            ('ALIGN', (1,0), (1,-1), 'RIGHT'),
            ('BOTTOMPADDING', (0,0), (-1,-1), 3),
            ('LINEABOVE', (0,4), (-1,4), 0.5, colors.black),
        ]))
        story.append(t)

        # Hasil tangkap
        if tangkap_list.exists():
            story.append(Paragraph('Hasil Tangkap', h2))
            rows = [['Jenis Ikan', 'Berat (kg)', 'Kondisi']]
            for ht in tangkap_list:
                rows.append([ht.jenis_ikan.nama, f"{float(ht.berat_kg):.1f}", ht.get_kondisi_display()])
            t = Table(rows, colWidths=[8*cm, 4*cm, 4*cm])
            t.setStyle(tbl_style)
            story.append(t)

        # Penjualan
        if jual_list.exists():
            story.append(Paragraph('Penjualan', h2))
            rows = [['Jenis Ikan', 'Pembeli', 'Berat (kg)', 'Harga/kg', 'Total']]
            for p in jual_list:
                rows.append([
                    p.hasil_tangkap.jenis_ikan.nama,
                    p.pembeli.nama,
                    f"{float(p.berat_terjual):.1f}",
                    rp(p.harga_per_kg),
                    rp(float(p.berat_terjual) * float(p.harga_per_kg)),
                ])
            t = Table(rows, colWidths=[4*cm, 3.5*cm, 2.5*cm, 2.5*cm, 3.5*cm])
            t.setStyle(tbl_style)
            story.append(t)

        # Bagi hasil
        if bagi_list.exists():
            story.append(Paragraph('Bagi Hasil ABK', h2))
            rows = [['ABK', 'Nominal', 'Status']]
            for bh in bagi_list:
                rows.append([bh.abk.nama, rp(bh.nominal),
                             'Lunas' if bh.sudah_dibayar else 'Belum'])
            t = Table(rows, colWidths=[7*cm, 5*cm, 4*cm])
            t.setStyle(tbl_style)
            story.append(t)

        doc.build(story)
        return response
